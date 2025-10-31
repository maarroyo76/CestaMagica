from io import BytesIO
import os
import zipfile
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from django.db import transaction
from django.core.paginator import Paginator
from django.core.files.images import ImageFile
from django.utils.dateparse import parse_date
import openpyxl
from Cart.cart import Cart

from .models import Producto, Categoria, Marca, userProfile
from .decorators import role_required
from pedido.models import Pedido, DetallePedido
from .forms import RegistroForm, ProductoForm 

def home(request):
    productos_destacados = Producto.objects.select_related('categoria', 'marca').filter(destacado=True)[:8]
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
        'productos_destacados': productos_destacados
    }
    return render(request, 'CestaMagica/home.html', context)

def productos(request):
    perfil = request.session.get('perfil')
    productos = Producto.objects.select_related('categoria', 'marca').all()
    
    categorias = Categoria.objects.all().order_by('nombre')
    marcas = Marca.objects.all().order_by('nombre')

    search_query = request.GET.get('search', '')
    marca = request.GET.get('marca', '')
    categoria_id = request.GET.get('categoria', '')
    orden = request.GET.get('orden', 'nombre')

    if search_query:
        productos = productos.filter(nombre__icontains=search_query)
    if marca:
        productos = productos.filter(marca__nombre__icontains=marca)
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)

    if orden == 'nombre_desc':
        productos = productos.order_by('-nombre')
    elif orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('nombre')

    paginator = Paginator(productos, 16)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'perfil': perfil,
        'productos': page_obj,
        'categorias': categorias,
        'selected_categoria': categoria_id,
        'marcas': marcas,
        'selected_marca': marca,
        'search_query': search_query,
        'marca': marca,
        'categoria_id': categoria_id,
        'orden': orden,
        'page_obj': page_obj,
    }
    return render(request, 'CestaMagica/productos.html', context)

def search_autocomplete(request):
    """
    Vista para la búsqueda en vivo del navbar.
    Devuelve resultados de productos en formato JSON.
    """
    if 'term' in request.GET:
        term = request.GET.get('term')
        qs = Producto.objects.filter(nombre__icontains=term)
        
        qs = qs[:5]
        
        productos = []
        for producto in qs:
            productos.append({
                'id': producto.id,
                'nombre': producto.nombre,
                'url': reverse('detalle_producto', args=[producto.id]),
                'imagen': producto.imagen.url,
                'precio': producto.precio,
            })
            
        return JsonResponse(productos, safe=False)
    return JsonResponse([], safe=False)

@role_required('admin', 'staff')
def gestion(request):
    perfil = request.session.get('perfil')

    productos = Producto.objects.select_related('categoria', 'marca').all()
    
    categorias = Categoria.objects.all().order_by('nombre')
    marcas = Marca.objects.all().order_by('nombre')
    pedidos = Pedido.objects.all().order_by('-fecha')
    
    search_query = request.GET.get('search', '')
    marca_id = request.GET.get('marca', '')
    categoria_id = request.GET.get('categoria', '')

    if search_query:
        productos = productos.filter(nombre__icontains=search_query)
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    estado_pedido = request.GET.get('estado_pedido')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    if estado_pedido:
        pedidos = pedidos.filter(estado=estado_pedido)
    if fecha_desde:
        fecha_desde_dt = parse_date(fecha_desde)
        if fecha_desde_dt:
            pedidos = pedidos.filter(fecha__date__gte=fecha_desde_dt)
    if fecha_hasta:
        fecha_hasta_dt = parse_date(fecha_hasta)
        if fecha_hasta_dt:
            pedidos = pedidos.filter(fecha__date__lte=fecha_hasta_dt)
    
    context = {
        'perfil': perfil,
        'productos': productos,
        'categorias': categorias,
        'marcas': marcas,
        'pedidos': pedidos,
    }
    return render(request, 'CestaMagica/Gestion/gestion.html', context)


@require_POST
@role_required('admin', 'staff')
def agregar_marca(request):
    nombre = request.POST.get('nombre')
    if nombre:
        Marca.objects.create(nombre=nombre)
        messages.success(request, "Marca agregada correctamente.")
    return redirect('gestion')

@role_required('admin', 'staff')
def editar_marca(request, marca_id):
    marca = get_object_or_404(Marca, id=marca_id)
    nombre = request.POST.get('nombre', '').strip()

    if not nombre:
        return HttpResponseBadRequest("El nombre de la marca no puede estar vacío.")
    
    marca.nombre = nombre
    marca.save()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok'})
    
    messages.success(request, "Marca actualizada correctamente.")
    return redirect('gestion')

@role_required('admin', 'staff')
def eliminar_marca(request, marca_id):
    marca = get_object_or_404(Marca, id=marca_id)
    marca.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok'})

    messages.success(request, "Marca eliminada.")
    return redirect('gestion')

@require_POST
@role_required('admin', 'staff')
def agregar_categoria(request):
    nombre = request.POST.get('nombre')
    if nombre:
        Categoria.objects.create(nombre=nombre)
        messages.success(request, "Categoría agregada correctamente.")
    return redirect('gestion')

@role_required('admin', 'staff')
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    nombre = request.POST.get('nombre', '').strip()

    if not nombre:
        return HttpResponseBadRequest("Nombre requerido")

    categoria.nombre = nombre
    categoria.save()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok'})
    
    messages.success(request, "Categoría actualizada.")
    return redirect('gestion')

@role_required('admin', 'staff')
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    categoria.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok'})

    messages.success(request, "Categoría eliminada.")
    return redirect('gestion')

def contacto(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        email = request.POST.get('email')
        mensaje = request.POST.get('mensaje')
        print(nombre, email, mensaje)
        messages.success(request, 'Mensaje enviado correctamente')
        return redirect('contacto')
    
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
    }
    return render(request, 'CestaMagica/contacto.html', context)

@require_POST
@role_required('admin', 'staff')
def actualizar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    nuevo_estado = request.POST.get('estado')

    estados_validos = [codigo for codigo, nombre in Pedido.ESTADOS]

    if nuevo_estado in estados_validos:
        pedido.estado = nuevo_estado
        pedido.save()
        messages.success(request, "Estado del pedido actualizado.")
    else:
        messages.error(request, "Estado inválido.")
    return redirect('gestion')

@role_required('admin', 'staff')
def agregar_producto(request):
    perfil = request.session.get('perfil')
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto agregado correctamente')
            return redirect('gestion')
    else:
        form = ProductoForm()

    context = {
        'perfil': perfil,
        'form': form
    }
    return render(request, 'CestaMagica/Gestion/agregar.html', context)

@role_required('admin', 'staff')
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    perfil = request.session.get('perfil')

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto editado correctamente')
            return redirect('gestion')
    else:
        form = ProductoForm(instance=producto)

    context = {
        'perfil': perfil,
        'producto': producto,
        'form': form
    }
    return render(request, 'CestaMagica/Gestion/editar.html', context)

@role_required('admin', 'staff')
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.delete()
    messages.success(request, 'Producto eliminado correctamente')
    return redirect('gestion')

def detalle_producto(request, id):
    producto = get_object_or_404(Producto.objects.select_related('categoria', 'marca'), id=id)
    referer = request.META.get('HTTP_REFERER', reverse('productos'))
    perfil = request.session.get('perfil')

    context = {
        'producto': producto,
        'perfil': perfil,
        'referer': referer,
    }
    return render(request, 'CestaMagica/detalle.html', context)

def registro(request):
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            newUser = User.objects.create_user(
                username=data['username'],
                first_name=data['nombre'],
                last_name=data['apellido'],
                email=data['email'],
                password=data['password']
            )
            userProfile.objects.create(
                user=newUser,
                telefono='+56 9 ' + data['telefono'],
                role='cliente',
            )
            
            login(request, newUser)
            messages.success(request, '¡Registro exitoso! Bienvenido.')
            return redirect("home")
    else:
        form = RegistroForm()
        
    return render(request, 'CestaMagica/auth/registro.html', {'form': form})

def inicio_sesion(request):
    if request.method == 'POST':
        username = request.POST.get('user')
        clave = request.POST.get('pass')

        user = authenticate(request, username=username, password=clave)

        if user is not None:
            profile = userProfile.objects.get(user=user)
            request.session['perfil'] = profile.role

            login(request, user)
            if profile.role == 'admin' or profile.role == 'staff':
                return redirect('gestion')
            else:
                return redirect('home')
        else:
            return render(request, 'CestaMagica/auth/inicio_sesion.html', {'error': 'Usuario o contraseña incorrectos'})

    return render(request, 'CestaMagica/auth/inicio_sesion.html')

@role_required('admin', 'staff', 'cliente')
def cerrar_sesion(request):
    logout(request)
    return redirect('home')

@role_required('admin', 'staff', 'cliente')
def perfil(request):
    user = request.user
    try:
        profile = userProfile.objects.get(user=user)
    except userProfile.DoesNotExist:
        profile = None

    perfil_session = request.session.get('perfil')
    context = {
        'user': user,
        'perfil': perfil_session,
        'profile': profile
    }
    return render(request, 'CestaMagica/perfil.html', context)

def error_404_view(request, exception):
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
    }
    return render(request, 'CestaMagica/404.html', status=404, context=context)

def error_500_view(request):
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
    }
    return render(request, 'CestaMagica/500.html', status=500, context=context)

def error_403_view(request, exception):
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
    }
    return render(request, 'CestaMagica/403.html', status=403, context=context)

def error_400_view(request, exception):
    perfil = request.session.get('perfil')
    context = {
        'perfil': perfil,
    }
    return render(request, 'CestaMagica/400.html', status=400, context=context)


@role_required('admin', 'staff', 'cliente')
def historial_pedidos(request):
    perfil = request.session.get('perfil')
    pedidos = Pedido.objects.filter(
        usuario=request.user
    ).prefetch_related(
        'items__producto').order_by('-fecha')

    context = {
        'perfil': perfil,
        'pedidos': pedidos
    }
    return render(request, 'CestaMagica/historial_pedidos.html', context)


@role_required('admin', 'staff', 'cliente')
@require_POST
def recomprar_pedido(request, pedido_id):
    cart = Cart(request)
    
    pedido_original = get_object_or_404(Pedido, id=pedido_id, usuario=request.user)

    productos_agregados = []
    productos_omitidos = []

    for item in pedido_original.items.all():
        producto = item.producto
        cantidad = item.cantidad

        if producto.is_available and producto.stock_vendible >= cantidad:
            cart.add(producto, cantidad)
            productos_agregados.append(producto.nombre)
        else:
            productos_omitidos.append(producto.nombre)
    
    if productos_agregados:
        messages.success(request, f"Se agregaron al carrito los productos del pedido {pedido_original.codigo_pedido}.")
    
    if productos_omitidos:
        nombres_omitidos = ", ".join(productos_omitidos)
        messages.warning(request, f"No se pudieron agregar: {nombres_omitidos} (sin stock suficiente).")

    return redirect('cart:cart_detail')


@role_required('admin', 'staff')
def carga_masiva(request):
    perfil = request.session.get('perfil')
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        zip_file = request.FILES.get('zip_file')

        if not excel_file:
            messages.error(request, "Debes subir al menos un archivo Excel.")
            return redirect('carga_masiva')

        imagenes_zip = {}
        if zip_file:
            try:
                with zipfile.ZipFile(zip_file, 'r') as zf:
                    for filename in zf.namelist():
                        if not filename.startswith('__MACOSX'):
                            nombre_base = os.path.basename(filename)
                            imagenes_zip[nombre_base] = BytesIO(zf.read(filename))
            except zipfile.BadZipFile:
                messages.error(request, "El archivo ZIP está dañado.")
                return redirect('carga_masiva')

        productos_creados = 0
        productos_actualizados = 0
        errores = [] # Para reportar IDs no encontrados

        try:
            with transaction.atomic():
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                headers = [str(cell.value) for cell in sheet[1]]
                headers_lower = [h.lower().strip() for h in headers]
                
                try:
                    # --- 1. AÑADIMOS 'id' A LAS CABECERAS ---
                    idx_id = headers_lower.index('id') 
                    idx_id_tienda = headers_lower.index('id_tienda')
                    idx_nombre = headers_lower.index('nombre')
                    idx_desc = headers_lower.index('descripcion')
                    idx_precio = headers_lower.index('precio')
                    idx_stock = headers_lower.index('stock')
                    idx_cat = headers_lower.index('categoria')
                    idx_marca = headers_lower.index('marca')
                except ValueError as e:
                    messages.error(request, f"Falta la columna requerida en el Excel: {e}")
                    return redirect('carga_masiva')

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # --- 2. LEEMOS EL ID DE LA BASE DE DATOS ---
                    db_id = row[idx_id] 
                    nombre = row[idx_nombre]
                    
                    if not nombre: # Si no hay nombre, saltamos la fila
                        continue

                    # Leemos el resto de los datos
                    id_tienda = row[idx_id_tienda]
                    descripcion = row[idx_desc]
                    precio = int(row[idx_precio])
                    stock = int(row[idx_stock])
                    categoria_nombre = row[idx_cat]
                    marca_nombre = row[idx_marca]

                    categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
                    marca, _ = Marca.objects.get_or_create(nombre=marca_nombre)

                    # --- 3. LÓGICA DE ACTUALIZAR O CREAR BASADA EN 'db_id' ---
                    
                    if db_id:
                        # RUTA DE ACTUALIZACIÓN (el ID existe)
                        try:
                            producto = Producto.objects.get(id=db_id)
                            
                            # Actualizamos todos los campos
                            producto.id_tienda = id_tienda
                            producto.nombre = nombre
                            producto.descripcion = descripcion
                            producto.precio = precio
                            producto.stock = stock
                            producto.categoria = categoria
                            producto.marca = marca
                            producto.save() # Guarda todos los cambios
                            
                            productos_actualizados += 1
                        except Producto.DoesNotExist:
                            errores.append(f"El ID {db_id} ('{nombre}') no existe. Se ignoró esta fila.")
                    
                    else:
                        # RUTA DE CREACIÓN (el ID está en blanco)
                        # (Opcional: puedes añadir un chequeo de duplicados aquí si quieres)
                        
                        producto = Producto.objects.create(
                            id_tienda=id_tienda,
                            nombre=nombre,
                            descripcion=descripcion,
                            precio=precio,
                            stock=stock,
                            categoria=categoria,
                            marca=marca
                        )
                        productos_creados += 1
                        
                        # Asignar imagen (solo para productos nuevos)
                        imagen_encontrada = None
                        for extension in ['.png', '.jpg', '.jpeg', '.webp']:
                            if (nombre + extension) in imagenes_zip:
                                imagen_encontrada = imagenes_zip[nombre + extension]
                                break
                        
                        if imagen_encontrada:
                            img_file = ImageFile(imagen_encontrada, name=f"{nombre}.png")
                            producto.imagen = img_file
                            producto.save(update_fields=['imagen'])
                        
            messages.success(request, f"Carga completada: {productos_creados} productos creados, {productos_actualizados} productos actualizados.")
            if errores:
                messages.warning(request, "Se encontraron problemas: " + " ".join(errores))
            
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}. No se guardó ningún cambio.")
        
        return redirect('gestion')

    return render(request, 'CestaMagica/Gestion/carga_masiva.html', {'perfil': perfil})


@role_required('admin', 'staff')
def descargar_plantilla_productos(request): # Renombré la función
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # --- AÑADIMOS 'id' A LA PLANTILLA ---
    headers = ['id_web', 'id_tienda', 'nombre', 'descripcion', 'precio', 'stock', 'categoria', 'marca']
    ws.append(headers)
    
    # Marcamos la columna 'id' como importante
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    ws['A1'].font = openpyxl.styles.Font(bold=True, color="FF0000") # ID en rojo

    productos = Producto.objects.select_related('categoria', 'marca').all().order_by('nombre')

    for producto in productos:
        ws.append([
            producto.id, # 👈 AÑADIDO
            producto.id_tienda,
            producto.nombre,
            producto.descripcion,
            producto.precio,
            producto.stock,
            producto.categoria.nombre,
            producto.marca.nombre
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos_actual.xlsx"'
    
    return response